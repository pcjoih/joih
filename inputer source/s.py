from tkinter import *
import openpyxl
import random
import pyperclip
import keyboard
from time import sleep

wb = openpyxl.load_workbook("last.xlsx")
sh1 = wb.active
sh1 = wb["공통"]
sh2 = wb.active
sh2 = wb["SPR"]
sh3 = wb.active
sh3 = wb["OBS"]
sh4 = wb.active
sh4 = wb["AD"]
sh5 = wb.active
sh5 = wb["Dep"]

root = Tk()
root.title("기록지 입력기")
root.geometry("400x350-0+0")
root.resizable(False, False)

def p1():
    text.delete(1.0, END)
    state.delete(1.0, END)
    ran1 = sh1.cell(row = random.randrange(1, len(sh1["A"]) + 1), column = 1).value + "\n"
    text.insert(1.0, ran1)
    pyperclip.copy(ran1)
    state.insert(1.0, "공통")
    
def p2():
    text.delete(1.0, END)
    state.delete(1.0, END)
    ran1 = sh2.cell(row = random.randrange(1, len(sh2["A"]) + 1), column = 1).value + "\n"
    text.insert(1.0, ran1)
    pyperclip.copy(ran1)
    state.insert(1.0, "SPR")

def p3():
    text.delete(1.0, END)
    state.delete(1.0, END)
    ran1 = sh3.cell(row = random.randrange(1, len(sh3["A"]) + 1), column = 1).value + "\n"
    text.insert(1.0, ran1)
    pyperclip.copy(ran1)
    state.insert(1.0, "OBS")

def p4():
    text.delete(1.0, END)
    state.delete(1.0, END)
    ran1 = sh4.cell(row = random.randrange(1, len(sh4["A"]) + 1), column = 1).value + "\n"
    text.insert(1.0, ran1)
    pyperclip.copy(ran1)
    state.insert(1.0, "AD")

def p5():
    text.delete(1.0, END)
    state.delete(1.0, END)
    ran1 = sh5.cell(row = random.randrange(1, len(sh5["A"]) + 1), column = 1).value + "\n"
    text.insert(1.0, ran1)
    pyperclip.copy(ran1)
    state.insert(1.0, "Dep")

def p6():
    text.delete(1.0, END)
    state.delete(1.0, END)
    pyperclip.copy("")
    state.insert(1.0, "")

def kt(Event):
    s = state.get(1.0, END)
    if s.strip() in "공통":
        sleep(0.1)
        text.delete(1.0, END)
        ran1 = sh1.cell(row = random.randrange(1, len(sh1["A"]) + 1), column = 1).value + "\n"
        text.insert(1.0, ran1)
        pyperclip.copy(ran1)
    elif s.strip() in "SPR":
        sleep(0.1)
        text.delete(1.0, END)
        ran1 = sh2.cell(row = random.randrange(1, len(sh2["A"]) + 1), column = 1).value + "\n"
        text.insert(1.0, ran1)
        pyperclip.copy(ran1)
    elif s.strip() in "OBS":
        sleep(0.1)
        text.delete(1.0, END)
        ran1 = sh3.cell(row = random.randrange(1, len(sh3["A"]) + 1), column = 1).value + "\n"
        text.insert(1.0, ran1)
        pyperclip.copy(ran1)
    elif s.strip() in "AD":
        sleep(0.1)
        text.delete(1.0, END)
        ran1 = sh4.cell(row = random.randrange(1, len(sh4["A"]) + 1), column = 1).value + "\n"
        text.insert(1.0, ran1)
        pyperclip.copy(ran1)
    elif s.strip() in "Dep":
        sleep(0.1)
        text.delete(1.0, END)
        ran1 = sh5.cell(row = random.randrange(1, len(sh5["A"]) + 1), column = 1).value + "\n"
        text.insert(1.0, ran1)
        pyperclip.copy(ran1)

b1 = Button(root, text="공통", command=p1)
b2 = Button(root, text="SPR", command=p2)
b3 = Button(root, text="OBS", command=p3)
b4 = Button(root, text="AD", command=p4)
b5 = Button(root, text="Dep", command=p5)
b6 = Button(root, text="Clear", command=p6)
text = Text(root)
state = Text(root)

b1.place(x=10, y=10, width=70, height=30)
b2.place(x=10, y=50, width=70, height=30)
b3.place(x=10, y=90, width=70, height=30)
b4.place(x=10, y=130, width=70, height=30)
b5.place(x=10, y=170, width=70, height=30)
b6.place(x=10, y=210, width=70, height=30)
text.place(x=120, y=10, width=260, height=320)
state.place(x=10, y=300, width=70, height=30)

keyboard.add_hotkey("ctrl+v", kt, args=([""]))

wb.close()
root.mainloop()